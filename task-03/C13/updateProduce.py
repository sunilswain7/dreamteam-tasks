#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

# Load the workbook and select the relevant sheet
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row + 1):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value  # Get the produce name in column A
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]  # Update the price in column B

# Save the workbook to a new file
wb.save('updatedProduceSales.xlsx')
